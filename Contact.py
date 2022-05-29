from operator import contains
import os
import re
import openpyxl
from tqdm import tqdm
from time import sleep

xsls_path = os.getcwd() + "\\Output.xlsx"
read_path_1 = os.getcwd() + "\\A表.xlsx"
read_path_2 = os.getcwd() + "\\B表.xlsx"


def write_excel_xlsx(workbook, origin, result, index):
    sheet = workbook.active
    printInfo = ["", "", "", "", "", "", "", ""]
    if index == 0:
        printInfo = ["序号", "ERP物料编码", "ERP物料说明", "ERP单位",
                     "物资名称", "规格型号/订货号", "品牌／生产厂／产地", "单位"]
        index += 1
    elif result:
        printInfo = [str(index-1), str(result[0].internal_value), result[1].internal_value, result[2].internal_value,
                     origin[4].internal_value, origin[5].internal_value, origin[6].internal_value, origin[7].internal_value]
    elif origin:
        printInfo = [str(index-1), str(origin[1].internal_value), origin[2].internal_value, origin[3].internal_value,
                     origin[4].internal_value, origin[5].internal_value, origin[6].internal_value, origin[7].internal_value]

    col = 1

    for info in printInfo:
        if info != "None":
            sheet.cell(row=index, column=col, value=info)
        else:
            sheet.cell(row=index, column=col, value="")
        col += 1


def check_contains_str(strCon, type_name, info_array):
    for info in info_array:
        id = info[0].internal_value
        content = info[1].internal_value
        unit = info[2].internal_value
        if content.find(strCon) != -1 and content.find(type_name) != -1:
            return info


def check_contains_str_no_space(strCon, type_name, info_table):
    content = strCon.replace(" ", "")
    for key in info_table.keys():
        if key.find(content) != -1 and key.find(type_name) != -1:
            return info_table[key]


def check_contains_str_last(type_name, info_table):
    for key in info_table.keys():
        if key.find(type_name) != -1:
            return info_table[key]


def read_excel_xlsx():
    marched_row_str = []
    b_table = []
    b_table_no_space = {}
    workbook = openpyxl.load_workbook(read_path_1)
    sheet = workbook.active
    workbook2 = openpyxl.load_workbook(read_path_2)
    sheet2 = workbook2.active
    # for row in sheet.rows:
    for row in sheet.iter_rows(min_col=1, max_col=3):
        if row[0].internal_value and row[1].internal_value:
            if row[0].internal_value.isnumeric():
                b_table.append(row)
                no_space_str = row[1].internal_value.replace(" ", "")
                b_table_no_space[no_space_str] = row

    cnt = 0
    workbook_output = openpyxl.Workbook()
    write_excel_xlsx(workbook_output, None, None, 0)
    pbar = tqdm(total=sheet2.max_row)
    for row in sheet2.iter_rows(min_col=1, max_col=8):
        typename = row[4].internal_value
        strCont = str(row[5].internal_value)
        cnt += 1
        pbar.update(1)
        if typename and strCont:
            result = check_contains_str(strCont, typename, b_table)
            if not result:
                result = check_contains_str_no_space(
                    strCont, typename, b_table_no_space)
            if result:
                marched_row_str.append(strCont)
                
            write_excel_xlsx(workbook_output, row, result, cnt)

    pbar.close()
    workbook_output.save(xsls_path)
    workbook_output.close()


def main():
    read_excel_xlsx()
    print("hello")

main()
